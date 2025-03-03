import pandas as pd
import numpy as np
from prophet import Prophet
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta

# IFRS17 and RBC2 Functions
def create_sheet(wb, sheet_name, headers, data):
    """Create and fill a worksheet"""
    sheet = wb.create_sheet(title=sheet_name)
    sheet.append(headers)
    for row in dataframe_to_rows(data, index=False, header=False):
        sheet.append(row)
    return sheet

def calculate_bel(cash_flows, discount_rate, periods):
    """Calculate Best Estimate Liability (IFRS17)"""
    return sum(cf / (1 + discount_rate)**t for t, cf in enumerate(cash_flows[:periods], 1))

def calculate_risk_margin(impact, probability, risk_weight):
    """Calculate risk margin based on risk factors"""
    return impact * probability * risk_weight

def calculate_csm(fulfilment_cashflows, ra, acquisition_costs):
    """Calculate Contractual Service Margin"""
    return max(0, -(fulfilment_cashflows + ra - acquisition_costs))

def calculate_rbc_components(assets, liabilities, premiums):
    """Calculate RBC2 risk components"""
    c1 = assets * 0.04  # Credit risk
    c2 = liabilities * 0.05  # Insurance risk
    c3 = assets * 0.03  # Market risk
    c4 = premiums * 0.02  # Operational risk
    return {'C1': c1, 'C2': c2, 'C3': c3, 'C4': c4}

def calculate_rbc_ratio(available_capital, required_capital):
    """Calculate RBC2 ratio"""
    return (available_capital / required_capital) * 100

# Load historical data
historical_data = {
    'ds': pd.date_range(start='2023-01-01', periods=12, freq='M'),
    'y': [1000000, 1020000, 1040000, 1065000, 1080000, 1100000,
          1120000, 1140000, 1160000, 1180000, 1200000, 1220000]
}
df = pd.DataFrame(historical_data)

# Prophet model setup and forecast
model = Prophet(yearly_seasonality=True)
model.fit(df)
future = model.make_future_dataframe(periods=12, freq='M')
forecast = model.predict(future)

# Create workbook
workbook = Workbook()

# 1. Inputs sheet - ALM assumptions
inputs_headers = ['Parameter', 'Value']
inputs_data = pd.DataFrame({
    'Parameter': [
        'Initial Assets', 'Discount Rate', 'Risk-Free Rate', 
        'Credit Spread', 'Liability Duration', 'Target Funding Ratio',
        'IFRS17 Risk Adjustment CoC', 'RBC2 Target Ratio',
        'Acquisition Costs', 'Coverage Period'
    ],
    'Value': [
        1000000, 0.03, 0.02, 0.01, 10, 1.10,
        0.06, 150, 50000, 12
    ]
})
create_sheet(workbook, 'Assumptions', inputs_headers, inputs_data)

# 2. IFRS17 Technical Provisions sheet
tech_prov_headers = [
    'Date', 'BEL', 'Risk Adjustment', 'CSM', 
    'Loss Component', 'Coverage Units', 'Total Technical Provisions'
]
tech_prov_data = pd.DataFrame()
tech_prov_data['Date'] = forecast['ds']
tech_prov_data['BEL'] = calculate_bel([50000]*12, 0.03, 12)
tech_prov_data['Risk Adjustment'] = tech_prov_data['BEL'] * 0.06
tech_prov_data['CSM'] = calculate_csm(
    tech_prov_data['BEL'].mean(),
    tech_prov_data['Risk Adjustment'].mean(),
    50000
)
tech_prov_data['Loss Component'] = 0  # Initialize with zero
tech_prov_data['Coverage Units'] = 1000  # Simplified assumption
tech_prov_data['Total Technical Provisions'] = (
    tech_prov_data['BEL'] + 
    tech_prov_data['Risk Adjustment'] + 
    tech_prov_data['CSM']
)
create_sheet(workbook, 'IFRS17_Technical_Provisions', tech_prov_headers, tech_prov_data)

# 3. Assets sheet
assets_headers = ['Date', 'Market Value', 'Expected Return', 'Duration', 'Credit Quality']
assets_data = forecast[['ds', 'yhat']].copy()
assets_data.columns = ['Date', 'Market Value']
assets_data['Expected Return'] = 0.05
assets_data['Duration'] = 8
assets_data['Credit Quality'] = 'AA'
create_sheet(workbook, 'Assets', assets_headers, assets_data)

# 4. RBC2 Capital Requirements sheet
rbc_components = calculate_rbc_components(
    assets_data['Market Value'].mean(),
    tech_prov_data['Total Technical Provisions'].mean(),
    50000
)

rbc2_headers = ['Risk Component', 'Capital Required']
rbc2_data = pd.DataFrame({
    'Risk Component': [
        'C1 (Asset Risk)', 'C2 (Insurance Risk)', 
        'C3 (Market Risk)', 'C4 (Operational Risk)',
        'Total Required Capital', 'Available Capital', 
        'RBC2 Ratio (%)'
    ],
    'Capital Required': [
        *rbc_components.values(),
        np.sqrt(sum([v**2 for v in rbc_components.values()])),
        assets_data['Market Value'].mean() - tech_prov_data['Total Technical Provisions'].mean(),
        0  # Will be calculated below
    ]
})
rbc2_data.loc[6, 'Capital Required'] = calculate_rbc_ratio(
    rbc2_data.loc[5, 'Capital Required'],
    rbc2_data.loc[4, 'Capital Required']
)
create_sheet(workbook, 'RBC2_Capital', rbc2_headers, rbc2_data)

# 5. Risk Analysis sheet
risk_headers = ['Risk Type', 'Impact', 'Probability', 'Risk Weight', 'Risk Margin']
risk_types = ['Market Risk', 'Credit Risk', 'Insurance Risk', 'Operational Risk']
risk_data = pd.DataFrame({
    'Risk Type': risk_types,
    'Impact': [5000000, 3000000, 2000000, 1000000],
    'Probability': [0.05, 0.03, 0.04, 0.02],
    'Risk Weight': [1.5, 1.3, 1.2, 1.1],
    'Risk Margin': 0.0
})
risk_data['Risk Margin'] = risk_data.apply(
    lambda x: calculate_risk_margin(x['Impact'], x['Probability'], x['Risk Weight']), 
    axis=1
)
create_sheet(workbook, 'Risk_Analysis', risk_headers, risk_data)

# 6. Stress Testing sheet
stress_headers = ['Scenario', 'Asset Impact', 'Liability Impact', 'Capital Impact', 'RBC2 Ratio Impact']
stress_scenarios = {
    'Base': [1.0, 1.0],
    'Equity -30%': [0.7, 1.1],
    'Interest Rate +200bps': [0.95, 0.9],
    'Credit Spread +100bps': [0.97, 1.05],
    'Insurance Loss +50%': [1.0, 1.5]
}

stress_results = []
base_assets = assets_data['Market Value'].mean()
base_liabilities = tech_prov_data['Total Technical Provisions'].mean()
base_capital = base_assets - base_liabilities

for scenario, factors in stress_scenarios.items():
    asset_factor, liability_factor = factors
    stressed_assets = base_assets * asset_factor
    stressed_liabilities = base_liabilities * liability_factor
    stressed_capital = stressed_assets - stressed_liabilities
    rbc2_impact = (stressed_capital / base_capital - 1) * 100
    stress_results.append([
        stressed_assets - base_assets,
        stressed_liabilities - base_liabilities,
        stressed_capital - base_capital,
        rbc2_impact
    ])

stress_data = pd.DataFrame(
    stress_results,
    columns=['Asset Impact', 'Liability Impact', 'Capital Impact', 'RBC2 Ratio Impact'],
    index=stress_scenarios.keys()
)
stress_data.index.name = 'Scenario'
create_sheet(workbook, 'Stress_Testing', stress_headers, stress_data.reset_index())

# 7. Dashboard
dashboard_headers = ['Metric', 'Current', 'Target', 'Status']
dashboard_data = pd.DataFrame({
    'Metric': [
        'RBC2 Ratio', 'IFRS17 CSM Ratio', 'Duration Gap',
        'Risk Margin Ratio', 'Capital Adequacy'
    ],
    'Current': [
        rbc2_data.loc[6, 'Capital Required'],
        tech_prov_data['CSM'].mean() / tech_prov_data['Total Technical Provisions'].mean() * 100,
        assets_data['Duration'].mean() - tech_prov_data['Coverage Units'].mean(),
        risk_data['Risk Margin'].sum() / tech_prov_data['Total Technical Provisions'].mean() * 100,
        (base_assets / base_liabilities) * 100
    ],
    'Target': [150, 15, 0, 5, 110],
    'Status': ''
})
dashboard_data['Status'] = dashboard_data.apply(
    lambda x: 'GREEN' if x['Current'] >= x['Target'] else 'RED', axis=1
)
create_sheet(workbook, 'Dashboard', dashboard_headers, dashboard_data)

# Save workbook and create visualization
excel_file = 'ALM_RBC2_IFRS17_Model.xlsx'
workbook.save(excel_file)
print(f"Enhanced ALM model with RBC2 and IFRS17 requirements saved to {excel_file}")

# Plot key metrics
plt.figure(figsize=(15, 5))
plt.subplot(1, 2, 1)
plt.plot(tech_prov_data['Date'], tech_prov_data['Total Technical Provisions'], label='Technical Provisions')
plt.plot(assets_data['Date'], assets_data['Market Value'], label='Assets')
plt.title('Assets vs Technical Provisions')
plt.legend()

plt.subplot(1, 2, 2)
plt.bar(risk_data['Risk Type'], risk_data['Risk Margin'])
plt.title('Risk Margins by Risk Type')
plt.xticks(rotation=45)
plt.tight_layout()
plt.show()